# -*- coding: utf-8 -*-

"""
This script is the manager for excel files.
it will handle the logic for the excel files.
"""

__author__ = "Juno Park"
__github__ = "https://github.com/junopark00"

import os
import openpyxl.utils
import sgtk
from sgtk.platform.qt import QtGui, QtCore
from sgtk import TankError

try:
    from PIL import Image
    import openpyxl
    import openpyxl_image_loader
except ImportError as e:
    raise TankError("This script requires the following packages: Pillow, openpyxl, openpyxl-image-loader")


# Set standard sgtk logger
logger = sgtk.platform.get_logger(__name__)


class ExcelManager:
    def __init__(self):
        self._current_dir = os.path.dirname(__file__)
        self._temp_dir = os.path.join(self._current_dir, ".temp_images")
        self.row_heights = {}
        self.column_widths = {}
        self.header_labels = []
        self.cell_values = {}
        self.loaded_images = []

    def load_excel(self, excel_path: str):
        """
        Load the excel file and update UI.
        
        Args:
            excel_path (str): path to the excel file
        """
        self.row_heights = self.get_row_heights(excel_path)
        self.column_widths = self.get_column_widths(excel_path)
        self.header_labels = self.get_header_labels(excel_path)
        self.cell_values = self.get_cell_values(excel_path)
        self.loaded_images = self.get_images(excel_path)

    def get_row_heights(self, excel_path: str) -> dict:
        """
        Get the row heights of the excel file.
        Exclude the header row.
        
        Args:
            excel_path (str): path to the excel file
        """
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        row_heights = {
            row : sheet.row_dimensions[row].height
            for row in range(2, sheet.max_row + 1) # Skip the header row
        }
        wb.close()
        
        return row_heights
    
    def get_column_widths(self, excel_path: str) -> dict:
        """
        Get the column widths of the excel file.
        
        Args:
            excel_path (str): path to the excel file
        """
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        column_widths = {
            column: sheet.column_dimensions[
                openpyxl.utils.get_column_letter(column)
                ].width
            for column in range(1, sheet.max_column + 1)
        }
        wb.close()
        
        return column_widths
    
    def get_header_labels(self, excel_path: str) -> list:
        """
        Get the header labels of the excel file.
        
        Args:
            excel_path (str): path to the excel file
        """
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        header_labels = [cell.value for cell in sheet[1]]
        wb.close()
        
        return header_labels
        
    def get_cell_values(self, excel_path: str) -> dict:
        """
        Get the cell values of the excel file.
        
        Args:
            excel_path (str): path to the excel file
        """
        items = {}
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            for j, value in enumerate(row):
                item = QtGui.QTableWidgetItem(
                    str(value)
                    ) if value is not None else QtGui.QTableWidgetItem("")
                items[(i, j)] = item
        wb.close()        
        
        return items
    
    def get_images(self, excel_path: str) -> list:
        """
        Get the images of the excel file.
        
        Args:
            excel_path (str): path to the excel file
        """
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        image_loader = openpyxl_image_loader.SheetImageLoader(sheet)
        
        images = []
        for row in sheet.iter_rows():
            for cell in row:
                if image_loader.image_in(cell.coordinate):
                    img_obj = image_loader.get(cell.coordinate)
                    images.append((cell.row, cell.column, img_obj))
        
        for img in images:
            row, column, img_obj = img
            img = Image.open(img_obj.fp)
            
            if not os.path.exists(self._temp_dir):
                os.makedirs(self._temp_dir)
                
            # Save the image to the temp directory
            img_path = os.path.join(self._temp_dir, f"{row}_{column}.png")    
            self.loaded_images.append(img_path)
            img.save(img_path)
        wb.close()
        
        return self.loaded_images

    def save_excel(
        self, 
        header_data: list, 
        cell_data: list, 
        excel_path: str, 
        version_up: bool
        ) -> str:
        """
        Save the excel file.

        Args:
            excel_path (str): path to the excel file
            version_up (bool): whether to version up the file
            
        Returns:
            str: path to the saved excel file
        """
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # append header data as the first row
        sheet.append(header_data)
        
        # append cell data
        for row in cell_data:
            sheet.append(row)
            
        # save images to the excel file
        for img_path in os.listdir(self._temp_dir):
            if img_path.endswith(".png"):
                row, column = os.path.splitext(img_path)[0].split("_")
                img = openpyxl.drawing.image.Image(os.path.join(self._temp_dir, img_path))
                
                img.width, img.height = 304, 171
                sheet.add_image(
                    img, 
                    f"{openpyxl.utils.get_column_letter(int(column))}{row}"
                    )
        
        # apply row heights and column widths
        for row, height in self.row_heights.items():
            sheet.row_dimensions[row].height = height
        for column, width in self.column_widths.items():
            sheet.column_dimensions[
                openpyxl.utils.get_column_letter(column)
                ].width = width
            
        # get new excel path if version up
        if version_up:
            save_path = self.version_up(excel_path)
        else:
            save_path = excel_path
            
        wb.save(save_path)
        wb.close()
        
        return save_path
            
    def version_up(self, excel_path: str) -> str:
        """
        Version up the excel file.
        
        Args:
            excel_path (str): path to the excel file
            
        Returns:
            str: path to the versioned up excel file
        """
        # get current file name and extension
        file_name = os.path.basename(excel_path)
        file_name, ext = os.path.splitext(file_name)
        
        # check if the file name has a underscore
        if "_" not in file_name:
            basename = file_name
            version = "0"
        else:
            basename, version = file_name.rsplit("_", 1)
        
        # check if the num is a digit
        if version.isdigit():
            version = int(version) + 1
        else:
            version = 1
            
        # create new file name
        new_file_name = f"{basename}_{version:02d}{ext}"
        excel_dir = os.path.dirname(excel_path)
        
        # check if the new file name exists, if so, version up again
        while os.path.exists(os.path.join(excel_dir, new_file_name)):
            version += 1
            new_file_name = f"{basename}_{version:02d}{ext}"
        
        # create new file path
        save_path = os.path.join(excel_dir, new_file_name)
        
        return save_path